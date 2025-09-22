import os
import zipfile
from openpyxl import load_workbook


CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIRECTORY = os.path.dirname(CURRENT_FILE)
TMP_DIR = os.path.join(CURRENT_DIRECTORY, 'Resources')

def test_xlsx_into_archive():

    archive_path = os.path.join(TMP_DIR,'archive_resources.zip')

    with zipfile.ZipFile(archive_path, 'r') as zip_ref:
        with zip_ref.open('file_xlsx.xlsx') as xlsx_file_in_zip:

                workbook = load_workbook(xlsx_file_in_zip)
                sheet = workbook.active

                all_xlsx = list(sheet.iter_rows(values_only=True))
                assert all_xlsx[0][0] == 'Январь'





